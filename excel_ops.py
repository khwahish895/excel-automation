import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill


def read_excel(file):
    return pd.read_excel(file)


def clean_data(df):
    df = df.dropna()
    df = df.drop_duplicates()
    return df


def save_excel(df, path):
    df.to_excel(path, index=False)


def format_excel(file_path):
    wb = load_workbook(file_path)
    ws = wb.active

    header_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")

    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill

    wb.save(file_path)
